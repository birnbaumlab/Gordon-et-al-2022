
# standard libraries
import os
import collections

# nonstandard libraries
import openpyxl


def publish_datasets_to_excel(data_dict,directory='./datasets'):

    """ Publish handled data to .csv """

    wb = openpyxl.Workbook() # open new workbook

    print(data_dict.keys())

    _write_settings(wb,data_dict['SETTINGS'])

    ws = wb.create_sheet(title='DATA')

    ws.append(('Dataset','Round','BC','Count'))

    for dataset_name,datasets in data_dict['DATA'].items():
        for rnd_index,dataset in enumerate(datasets):
            for bc,count in dataset.items():
                ws.append((dataset_name,rnd_index,bc,count))

    export_name = _unique_wb_save(wb,'dataset',directory)
    
    return export_name


def publish_reference_to_excel(data_dict,directory='./references'):

    """ Publish handled data to .xlsx """

    wb = openpyxl.Workbook() # open new workbook

    # settings sheet
    _write_settings(wb,data_dict['SETTINGS'])

    # data sheet
    ws = wb.create_sheet(title='DATA')

    ws.append(('Barcode','Consensus Domains','% Freq.','All Domain Sets'))

    for k,v in data_dict['DATA'].items():
        if v:
            domain_dict = collections.Counter(v)
            common_domain = domain_dict.most_common(1)
            print((k,','.join(common_domain[0][0]),100*float(common_domain[0][1])/len(v))+tuple((','.join(i) for i in v)))
            ws.append((k,','.join(common_domain[0][0]),100*float(common_domain[0][1])/len(v))+tuple((','.join(i) for i in v)))
        else:
            print((k,','.join(''),0))
            ws.append((k,','.join(''),0))

    export_name = _unique_wb_save(wb,'reference',directory)
    
    return export_name


def _write_settings(wb,settings):
    """ Settings writing """
    ws = wb.create_sheet(title='SETTINGS')

    for k,v in settings.items():
        if isinstance(v,dict):
            ws.append((k,'dict->'))
            for k2,v2 in v.items():
                if isinstance(v2,(tuple,list)):
                    ws.append(('->',k2)+tuple(v2))
                else:
                    ws.append(('->',k2,v2))
        else:
            ws.append((k,v))

def _unique_wb_save(wb,name,directory):

    # reference object
    for i in range(1,10000):

        reference_fname = os.path.join(directory,'{}_{}.xlsx'.format(name,str(i).zfill(4)))
        fname = '{}_{}.xlsx'.format(name,str(i).zfill(4))

        if os.path.isfile(reference_fname): continue
            
        # try removing starting blank sheet
        try: del wb['Sheet']
        except KeyError: pass

        wb.save(reference_fname) # save final workbook

        return fname



