
# standard libraries
from math import log
import collections

# nonstandard libraries
import openpyxl

#---------------------------------#

def write_counts(datasets_dict,**kwargs):

    """ Compares arbitrary number of sets of frequency dictionaries for enrichment """

    settings = {
            'reference':None,
            'workbook':None,
            'silent':False,
            }

    # update settings
    if 'reference' in kwargs: settings['reference'] = kwargs['reference']
    if 'workbook' in kwargs:
        settings['workbook'] = kwargs['workbook']
    if 'silent' in kwargs:
        settings['silent'] = kwargs['silent']

    reference = settings['reference']
    wb = settings['workbook']
    silent = settings['silent']

    # create new workbook if one doesn't exist
    if not wb: wb = openpyxl.Workbook() # open new workbook
    
    # initialize processed dataset 
    comparison_datasets = {}

    # iterate across dataset in datasets for particular selection
    for dataset_name,raw_datasets in datasets_dict.items():

        comparison_datasets[dataset_name] = {}

        # transform dataset
        if reference:
            # if we have a reference to translate barcodes...
            bc2domain = [(k,collections.Counter(v).most_common(1)) 
                               for k,v in reference.items()]
            bc2domain = dict([(k,v[0][0]) for k,v in bc2domain if len(v) > 0])

            datasets = [{} for _ in raw_datasets] 

            for i,raw_dataset in enumerate(raw_datasets):
                for bc,count in raw_dataset.items():
                    if bc in bc2domain:
                        try:
                            datasets[i][bc2domain[bc]] += count
                        except KeyError:
                            datasets[i][bc2domain[bc]]  = count
                    else:
                        continue 

        else:
            # otherwise keep dictionary as is
            datasets = raw_datasets

        # create new worksheet in workbook
        ws = wb.create_sheet(title='Counts - {}'.format(dataset_name))

        # create data header with merge features
        data_header = [['',]   + 2*(['' for _ in range(len(datasets))]),
                       ['ID',] + 2*(['Round {}'.format(i) for i in range(len(datasets))])]
        data_header[0][1]                 = 'Count'
        data_header[0][1 + len(datasets)] = 'Frequency'

        # quick checks on data size
        if len(datasets) == 0:
            raise ValueError('No datasets passed to comparison function!')
        if len(datasets) == 1:
            if not silent:
                print('Only one dataset passed for comparison, skipping comparison...')
            datasets.append({})

        # Get total sequences in pre/post
        seq_total = [sum(d.values()) for d in datasets]
        sample_total = len(datasets)
        enrichment_data = {}

        # get a list of all barcodes
        barcodes = set().union(*[dataset.keys() for dataset in datasets])

        data = data_header[:] # preinitialize

        for bc in barcodes:

            counts = [] # initialize count array

            # Check for everything in presort
            for dataset in datasets:
                # get precount
                try:
                    counts.append(dataset[bc])
                except KeyError:
                    counts.append(0)
            
            freqs   = [float(c)/s if s > 0 else 'N/A' for c,s in zip(counts,seq_total)]
            counts += freqs

            # add counts to specific dataset object
            comparison_datasets[dataset_name][bc] = freqs 
            
            # store value
            data.append([bc] + counts)
            
        for row in data:
            ws.append(row)

        ws.merge_cells(start_row=1, start_column=2, 
                end_row=1, end_column=1+len(datasets))
        ws.merge_cells(start_row=1, start_column=2+len(datasets), 
                end_row=1, end_column=1+2*len(datasets))

    return comparison_datasets

#---------------------------------#

"""
archived code

        # Get total sequences in pre/post
        seq_total = [sum(d.values()) for d in seq_dicts]
        sample_total = len(datasets)
        enrichment_data = {}

        for i in range(sample_total-1):
            for j in range(i+1,sample_total):

                data = data_header[:] # preinitialize

                # Check for everything in presort
                for seq in seq_dicts[i]:
                    # get precount
                    pre = float(seq_dicts[i][seq])
                    try:
                        post = float(seq_dicts[j][seq])
                        ratio = log((post*seq_total[i])/(pre*seq_total[j]),2)
                        data.append((seq,pre,post,pre/seq_total[i],post/seq_total[j],ratio))
                    except KeyError:
                        data.append((seq,pre,'N/A',pre/seq_total[i],'N/A','N/A'))

                # Add things only in postsort
                for seq in seq_dicts[j]:
                    try:
                        seq_dicts[i][seq]
                    except KeyError:
                        post = seq_dicts[j][seq]
                        data.append((seq,'N/A',post,'N/A',post/seq_total[j],'N/A'))

                # store value
                enrichment_data['Round {}->{}'.format(i,j)] = data        

"""




